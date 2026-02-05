import os, re, zipfile, shutil
from openpyxl import load_workbook

# מציאת קובץ ה-XLSM
xlsm_files = [f for f in os.listdir('.') if f.endswith('.xlsm')]
if not xlsm_files:
    raise RuntimeError("No XLSM file found")
XLSM = xlsm_files[0]

# ניקוי תיקיית images אם קיימת
if os.path.exists("images"):
    shutil.rmtree("images")
os.makedirs("images", exist_ok=True)

wb = load_workbook(XLSM, data_only=True)
sheet = wb["P"]

saved = 0

for img in sheet._images:
    try:
        # דילוג על תמונות לא נתמכות (WMF/OLE)
        if not hasattr(img, "_data"):
            continue

        anchor = img.anchor._from
        row = anchor.row + 1

        qid = sheet.cell(row=row, column=12).value
        if not isinstance(qid, str):
            continue

        m = re.search(r"(\\d+)", qid)
        if not m:
            continue

        qnum = m.group(1)
        out_path = f"images/q_{qnum}.png"

        with open(out_path, "wb") as f:
            f.write(img._data())

        saved += 1

    except Exception:
        continue

# יצירת index.html בסיסי לבדיקה
with open("index.html", "w", encoding="utf-8") as f:
    f.write(f"""
<!DOCTYPE html>
<html lang="he">
<head>
<meta charset="UTF-8">
<title>בדיקת חילוץ תמונות</title>
</head>
<body>
<h1>חולצו {saved} תמונות</h1>
<ul>
""")
    for name in sorted(os.listdir("images")):
        f.write(f"<li>{name}</li>")
    f.write("</ul></body></html>")

# יצירת ZIP
with zipfile.ZipFile("site.zip", "w") as z:
    z.write("index.html")
    for f in os.listdir("images"):
        z.write("images/" + f)
